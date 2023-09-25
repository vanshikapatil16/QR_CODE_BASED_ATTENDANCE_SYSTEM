import os
from openpyxl import load_workbook
from MyQR import myqr

folder_path = "3rd_year"  # Specify the folder path here

# Create the folder if it doesn't exist
os.makedirs(folder_path, exist_ok=True)

file_path = 'data.xlsx'  # Replace with the actual file path

# Load the workbook
workbook = load_workbook(file_path)

# Select the first sheet
sheet = workbook.active

# Read the data from the sheet
lines = []
for row in sheet.iter_rows(values_only=True):
    line = ' '.join(str(cell) for cell in row)
    lines.append(line)

for line in lines:
    if line:
        data = line.strip()
        save_name = f"{data}.png"
        save_path = os.path.join(folder_path, save_name)  # Join the folder path with the save_name

        version, level, qr_name = myqr.run (
            data,
            version=1,
            colorized=True,
            contrast=1.0,
            brightness=1.0,
            save_name=save_name,
            save_dir=folder_path
        )
        print(f"QR code generated for {data} and saved at {save_path}")
